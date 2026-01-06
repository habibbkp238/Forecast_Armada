import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

class ExportManager:
    """Handles data export functionality"""
    
    def __init__(self):
        self.brand_color = '2E7D32'  # Green
    
    def export_detailed_excel(self, forecast_df, historical_df=None, metadata=None):
        """Export detailed forecast to Excel with multiple sheets"""
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Detailed Forecast
            forecast_export = forecast_df.copy()
            
            # Select and order columns
            export_cols = ['date', 'company', 'origin', 'destination', 
                          'province', 'region', 'fleet_type', 'forecast_qty']
            
            # Ensure all columns exist
            for col in export_cols:
                if col not in forecast_export.columns:
                    forecast_export[col] = 'All'
            
            forecast_export = forecast_export[export_cols]
            forecast_export.to_excel(writer, sheet_name='Detailed Forecast', index=False)
            
            # Sheet 2: Summary by Date
            summary_by_date = forecast_df.groupby('date')['forecast_qty'].sum().reset_index()
            summary_by_date.columns = ['Date', 'Total Forecast']
            summary_by_date.to_excel(writer, sheet_name='Summary by Date', index=False)
            
            # Sheet 3: Summary by Fleet Type
            if 'fleet_type' in forecast_df.columns:
                summary_by_fleet = forecast_df.groupby('fleet_type')['forecast_qty'].sum().reset_index()
                summary_by_fleet.columns = ['Fleet Type', 'Total Forecast']
                summary_by_fleet = summary_by_fleet.sort_values('Total Forecast', ascending=False)
                summary_by_fleet.to_excel(writer, sheet_name='Summary by Fleet Type', index=False)
            
            # Sheet 4: Summary by Region
            if 'region' in forecast_df.columns and forecast_df['region'].nunique() > 1:
                summary_by_region = forecast_df.groupby('region')['forecast_qty'].sum().reset_index()
                summary_by_region.columns = ['Region', 'Total Forecast']
                summary_by_region = summary_by_region.sort_values('Total Forecast', ascending=False)
                summary_by_region.to_excel(writer, sheet_name='Summary by Region', index=False)
            
            # Sheet 5: Summary by Route
            if 'origin' in forecast_df.columns and 'destination' in forecast_df.columns:
                route_summary = forecast_df.groupby(['origin', 'destination'])['forecast_qty'].sum().reset_index()
                route_summary.columns = ['Origin', 'Destination', 'Total Forecast']
                route_summary = route_summary.sort_values('Total Forecast', ascending=False)
                route_summary.to_excel(writer, sheet_name='Summary by Route', index=False)
            
            # Sheet 6: Metadata
            if metadata:
                metadata_df = pd.DataFrame(list(metadata.items()), columns=['Property', 'Value'])
                metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
        
        # Format the workbook
        output.seek(0)
        workbook = openpyxl.load_workbook(output)
        
        # Format each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Header styling
            header_fill = PatternFill(start_color=self.brand_color, 
                                     end_color=self.brand_color, 
                                     fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
        
        # Save formatted workbook
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output
    
    def export_summary_excel(self, forecast_df, metadata=None):
        """Export summary forecast to Excel"""
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Summary by date
            summary_by_date = forecast_df.groupby('date')['forecast_qty'].sum().reset_index()
            summary_by_date.columns = ['Date', 'Total Forecast']
            summary_by_date.to_excel(writer, sheet_name='Summary', index=False)
            
            # Optional: Add fleet type breakdown
            if 'fleet_type' in forecast_df.columns and forecast_df['fleet_type'].nunique() > 1:
                fleet_pivot = forecast_df.pivot_table(
                    index='date',
                    columns='fleet_type',
                    values='forecast_qty',
                    aggfunc='sum',
                    fill_value=0
                )
                fleet_pivot.to_excel(writer, sheet_name='By Fleet Type')
            
            # Metadata
            if metadata:
                metadata_df = pd.DataFrame(list(metadata.items()), columns=['Property', 'Value'])
                metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
        
        # Format the workbook
        output.seek(0)
        workbook = openpyxl.load_workbook(output)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Header styling
            header_fill = PatternFill(start_color=self.brand_color, 
                                     end_color=self.brand_color, 
                                     fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output
    
    def export_detailed_csv(self, forecast_df):
        """Export detailed forecast to CSV"""
        
        export_cols = ['date', 'company', 'origin', 'destination', 
                      'province', 'region', 'fleet_type', 'forecast_qty']
        
        # Ensure all columns exist
        for col in export_cols:
            if col not in forecast_df.columns:
                forecast_df[col] = 'All'
        
        forecast_export = forecast_df[export_cols].copy()
        
        return forecast_export.to_csv(index=False).encode('utf-8')
    
    def export_summary_csv(self, forecast_df):
        """Export summary forecast to CSV"""
        
        summary_by_date = forecast_df.groupby('date')['forecast_qty'].sum().reset_index()
        summary_by_date.columns = ['date', 'total_forecast']
        
        return summary_by_date.to_csv(index=False).encode('utf-8')
    
    def create_metadata_dict(self, forecast_df, model_used, aggregation_level, 
                            horizon, freq, include_holidays):
        """Create metadata dictionary for export"""
        
        metadata = {
            'Export Date': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
            'Model Used': model_used,
            'Aggregation Level': aggregation_level,
            'Forecast Horizon': horizon,
            'Frequency': 'Daily' if freq == 'D' else 'Monthly',
            'Holidays Included': 'Yes' if include_holidays else 'No',
            'Total Forecast Periods': len(forecast_df['date'].unique()),
            'Total Series Forecasted': forecast_df['unique_id'].nunique() if 'unique_id' in forecast_df.columns else 1,
            'Total Forecasted Quantity': int(forecast_df['forecast_qty'].sum()),
            'Average Daily Forecast': int(forecast_df['forecast_qty'].mean())
        }
        
        return metadata
